"""基础文件读取：区分纯文本、多模态图片与 Office/PDF 文档文本提取。"""

from __future__ import annotations

import base64
import logging
import mimetypes
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from enum import Enum
from pathlib import Path

logger = logging.getLogger(__name__)

# DashScope 兼容模式：单条 data-uri 最大约 20 MiB（见 API 报错 20971520）
DASHSCOPE_MAX_MULTIMODAL_BYTES = 20 * 1024 * 1024

IMAGE_EXTENSIONS = frozenset(
    {".png", ".jpg", ".jpeg", ".gif", ".webp", ".heic", ".heif", ".bmp"}
)
DOCUMENT_EXTENSIONS = frozenset(
    {
        ".pdf",
        ".doc",
        ".docx",
        ".xls",
        ".xlsx",
        ".ppt",
        ".pptx",
    }
)
SHEET_EXTENSIONS = frozenset({".csv", ".tsv"})


class BasicFileCategory(str, Enum):
    TEXT = "text"
    IMAGE = "image"
    DOCUMENT = "document"


@dataclass(frozen=True)
class BasicFileContent:
    category: BasicFileCategory
    text: str | None = None
    base64_data: str | None = None
    mime_type: str | None = None


def categorize_file(path: Path | str) -> BasicFileCategory:
    ext = Path(path).suffix.lower()
    if ext in IMAGE_EXTENSIONS:
        return BasicFileCategory.IMAGE
    if ext in DOCUMENT_EXTENSIONS:
        return BasicFileCategory.DOCUMENT
    return BasicFileCategory.TEXT


def infer_resource_artifact_type(path: Path | str) -> str:
    """资源列表/预览用的 artifact_type。"""
    ext = Path(path).suffix.lower()
    if ext in IMAGE_EXTENSIONS:
        return "image"
    if ext in SHEET_EXTENSIONS or ext == ".xlsx":
        return "sheet"
    if ext in DOCUMENT_EXTENSIONS:
        return "document"
    from src.service.agent.artifacts import infer_artifact_type

    return infer_artifact_type(str(path))


def estimate_base64_decoded_bytes(base64_data: str) -> int:
    """估算 base64 解码后的字节数。"""
    padding = base64_data.count("=")
    return max(0, len(base64_data) * 3 // 4 - padding)


def is_multimodal_payload_too_large(
    *,
    raw_bytes: int | None = None,
    base64_data: str | None = None,
) -> bool:
    if raw_bytes is not None:
        return raw_bytes > DASHSCOPE_MAX_MULTIMODAL_BYTES
    if base64_data:
        return (
            estimate_base64_decoded_bytes(base64_data)
            > DASHSCOPE_MAX_MULTIMODAL_BYTES
        )
    return False


def format_multimodal_size_error(path: str, size_bytes: int) -> str:
    limit_mb = DASHSCOPE_MAX_MULTIMODAL_BYTES // (1024 * 1024)
    size_mb = size_bytes / (1024 * 1024)
    return (
        f"Error: 文件 {path} 约 {size_mb:.1f}MB，超过当前模型多模态上限 {limit_mb}MB。"
        f"请压缩图片后重新上传，或改用文本/PDF 等方式提供内容。"
    )


def read_basic_file(path: Path) -> BasicFileContent:
    category = categorize_file(path)
    if category == BasicFileCategory.IMAGE:
        raw = path.read_bytes()
        if is_multimodal_payload_too_large(raw_bytes=len(raw)):
            raise ValueError(format_multimodal_size_error(str(path), len(raw)))
        mime, _ = mimetypes.guess_type(path.name)
        return BasicFileContent(
            category=category,
            base64_data=base64.standard_b64encode(raw).decode("ascii"),
            mime_type=mime or "application/octet-stream",
        )
    if category == BasicFileCategory.DOCUMENT:
        extracted = extract_document_text(path)
        header = f"[已从 {path.name} 提取文本内容]\n\n"
        return BasicFileContent(category=category, text=header + extracted)
    return BasicFileContent(
        category=category, text=path.read_text(encoding="utf-8")
    )


def extract_document_text(path: Path) -> str:
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            return _extract_pdf(path)
        if ext == ".docx":
            return _extract_docx(path)
        if ext == ".xlsx":
            return _extract_xlsx(path)
        if ext == ".pptx":
            return _extract_pptx(path)
        if ext == ".doc":
            return _legacy_format_hint(".doc", ".docx")
        if ext == ".xls":
            return _legacy_format_hint(".xls", ".xlsx")
        if ext == ".ppt":
            return _legacy_format_hint(".ppt", ".pptx")
    except Exception as exc:
        logger.warning("文档解析失败 path=%s: %s", path, exc, exc_info=True)
        return f"[无法解析文档 {path.name}: {exc}]"
    return f"[不支持的文档格式: {ext}]"


def _legacy_format_hint(from_ext: str, to_ext: str) -> str:
    return f"[暂不支持 {from_ext} 格式，请转换为 {to_ext} 后重试]"


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        page_text = (page.extract_text() or "").strip()
        parts.append(f"## 第 {index} 页\n\n{page_text}")
    body = "\n\n".join(parts).strip()
    return body or "[PDF 未提取到文本（可能为扫描件）]"


def _extract_docx(path: Path) -> str:
    from docx import Document

    document = Document(str(path))
    paragraphs = [p.text.strip() for p in document.paragraphs if p.text.strip()]
    if not paragraphs:
        return "[Word 文档无可见文本段落]"
    return "\n\n".join(paragraphs)


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    sections: list[str] = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows: list[str] = []
            for row in worksheet.iter_rows(values_only=True):
                cells = ["" if value is None else str(value) for value in row]
                if any(cell.strip() for cell in cells):
                    rows.append("\t".join(cells))
            body = "\n".join(rows).strip()
            sections.append(
                f"## 工作表: {sheet_name}\n\n{body or '[空表]'}"
            )
    finally:
        workbook.close()
    return "\n\n".join(sections).strip()


def _extract_pptx(path: Path) -> str:
    sections: list[str] = []
    with zipfile.ZipFile(path) as archive:
        slide_names = sorted(
            name
            for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        for index, slide_name in enumerate(slide_names, start=1):
            root = ET.fromstring(archive.read(slide_name))
            namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
            texts = [
                node.text.strip()
                for node in root.findall(".//a:t", namespace)
                if node.text and node.text.strip()
            ]
            if texts:
                sections.append(f"## 幻灯片 {index}\n\n" + "\n".join(texts))
    return "\n\n".join(sections).strip() or "[PPT 未提取到文本]"
